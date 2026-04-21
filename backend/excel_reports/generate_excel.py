import openpyxl
import tkinter as tk
from tkinter import filedialog

def create_test_file():
    # Creamos una ventana para elegir ruta
    root = tk.Tk()
    root.withdraw()
    
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Guardar reporte como..."
    )

    if not path:
        print("Operación cancelada por el usuario.")
        return

    # Crear el libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Agregar datos
    ws['A1'] = "ID"
    ws['B1'] = "Date"
    ws['C1'] = "Status"
    ws['A2'] = 1
    ws['B2'] = "2026-04-18"
    ws['C2'] = "Processed"

    # Guardar en la ruta seleccionada
    try:
        wb.save(path)
        print(f"¡Éxito! Archivo guardado en: {path}")
    except Exception as e:
        print(f"Error al guardar: {e}")

if __name__ == "__main__":
    create_test_file()