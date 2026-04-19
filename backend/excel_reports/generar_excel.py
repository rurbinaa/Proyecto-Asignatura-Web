import openpyxl
import os

# Path where the file should be
# Use r'' to handle Windows backslashes properly
path = r'C:\Users\yessi\Documents\Proyecto-Asignatura-Web\backend\excel_reports\excel_templates\plantilla_corporativa.xlsx'

def crear_archivo_prueba():
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add some test data
    ws['A1'] = "ID"
    ws['B1'] = "Date"
    ws['C1'] = "Status"
    
    ws['A2'] = 1
    ws['B2'] = "2026-04-18"
    ws['C2'] = "Processed"

    # Save the file
    wb.save(path)
    print(f"File successfully generated at: {path}")

if __name__ == "__main__":
    crear_archivo_prueba()