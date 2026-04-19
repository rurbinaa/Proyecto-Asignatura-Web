import os
from openpyxl import load_workbook

class ExcelService:
    def __init__(self, template_filename):
        template_dir = os.path.join(os.path.dirname(__file__), 'excel_templates')
        self.template_path = os.path.join(template_dir, template_filename)
        self.wb = None

    def load_workbook(self):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"The file {self.template_path} was not found.")
        self.wb = load_workbook(filename=self.template_path)
        return self.wb

    def write_cell(self, sheet_name, cell, value):
        if self.wb is None:
            raise ValueError("The workbook has not been loaded. Call load_workbook() first.")
        ws = self.wb[sheet_name]
        ws[cell] = value

    def write_range(self, sheet_name, start_cell, data):
        """
        Write a range of data starting from start_cell.
        data must be a list of lists (rows).
        """
        if self.wb is None:
            raise ValueError("The workbook has not been loaded. Call load_workbook() first.")
        ws = self.wb[sheet_name]
        start_row, start_col = self._cell_to_row_col(start_cell)
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=value)

    def _cell_to_row_col(self, cell):
        """
        Convert a cell reference like 'A1' to row and column (1-based).
        """
        from openpyxl.utils import column_index_from_string, coordinate_to_tuple
        col, row = coordinate_to_tuple(cell)
        return row, col

    def save_workbook(self, output_path):
        if self.wb is None:
            raise ValueError("The workbook has not been loaded. Call load_workbook() first.")
        self.wb.save(output_path)

