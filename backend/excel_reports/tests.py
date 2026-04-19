from django.test import TestCase
from .services import ExcelService
from openpyxl.workbook.workbook import Workbook
import tempfile
import os
from django.test import Client
from django.contrib.auth.models import User
from quality_data.models import Color, DefectType
from media_data.models import InspectionData, RevisionDefect

class ExcelServiceTest(TestCase):
    def setUp(self):
        self.service = ExcelService('plantilla_corporativa.xlsx')

    def test_load_workbook_returns_instance(self):
        wb = self.service.load_workbook()
        self.assertIsInstance(wb, Workbook)

    def test_workbook_is_not_empty(self):
        wb = self.service.load_workbook()
        self.assertGreater(len(wb.sheetnames), 0)

    def test_write_cell_and_save(self):
        wb = self.service.load_workbook()
        sheet_name = wb.sheetnames[0]  # Obtener el nombre de la primera hoja
        self.service.write_cell(sheet_name, 'A1', 'Test Value')
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            self.service.save_workbook(tmp.name)
            self.assertTrue(os.path.exists(tmp.name))
            os.unlink(tmp.name)  # Eliminar el archivo temporal después de la prueba

    def test_write_range(self):
        wb = self.service.load_workbook()
        sheet_name = wb.sheetnames[0]
        data = [['Date', 'Value'], ['2026-04-18', 100]]
        self.service.write_range(sheet_name, 'A1', data)

        ws = wb[sheet_name]
        self.assertEqual(ws['A1'].value, 'Date')
        self.assertEqual(ws['B1'].value, 'Value')
        self.assertEqual(ws['A2'].value, '2026-04-18')
        self.assertEqual(ws['B2'].value, 100)


class ReportGenerationTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='123')
        self.color = Color.objects.create(name="Red", is_active=True)
        self.defect_type = DefectType.objects.create(name="Seam", is_active=True)

    def test_generate_report_without_parameters(self):
        response = self.client.get('/excel/reporte/')
        self.assertEqual(response.status_code, 400)
        self.assertIn("Missing start_date and end_date parameters", response.content.decode())

    def test_generate_report_invalid_date(self):
        response = self.client.get('/excel/reporte/?start_date=invalid&end_date=2026-04-20')
        self.assertEqual(response.status_code, 400)
        self.assertIn("Invalid date format", response.content.decode())

    def test_generate_report_no_data(self):
        response = self.client.get('/excel/reporte/?start_date=2026-01-01&end_date=2026-01-31')
        self.assertEqual(response.status_code, 404)
        self.assertIn("No data in the selected date range", response.content.decode())

    def test_generate_report_with_data(self):
        inspection = InspectionData.objects.create(
            inspector=self.user, color=self.color, lot="A1", style="Style1", size="M"
        )
        RevisionDefect.objects.create(
            inspection=inspection, defect_type=self.defect_type, defect_size="M", defect_count=2
        )
        
        response = self.client.get('/excel/reporte/?start_date=2026-01-01&end_date=2026-12-31')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment; filename="report_', response['Content-Disposition'])