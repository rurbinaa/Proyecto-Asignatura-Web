from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from copy import copy
import datetime as dt

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter, range_boundaries

from excel_importer.date_utils import (
    apply_charfield_iso_date_range,
    apply_datefield_date_range,
)
from excel_importer.sheet_configs import (
    CORPORATE_XLSX_CANONICAL_TEMPLATE_RELATIVE_PATH,
    CORPORATE_XLSX_EXPORT_CONFIG,
    CORPORATE_XLSX_PLACEHOLDER_TEMPLATE_RELATIVE_PATH,
)
from quality_data.models import Container, QualityQcFa, SecondsA4, SecondsGeneral


class EmptyCorporateXlsxDataError(Exception):
    """Raised when no reportable rows exist for the requested range."""


@dataclass
class CorporateXlsxArtifact:
    file_bytes: bytes
    filename: str


class CorporateXlsxReportService:
    MODEL_REGISTRY = {
        "QualityQcFa": QualityQcFa,
        "SecondsA4": SecondsA4,
        "SecondsGeneral": SecondsGeneral,
        "Container": Container,
    }

    def __init__(self, template_path=None):
        self.project_root = self._resolve_project_root()
        self.template_path = self._resolve_template_path(template_path)

    @staticmethod
    def _resolve_project_root():
        service_path = Path(__file__).resolve()
        candidate_roots = [service_path.parents[2], service_path.parents[1]]

        for candidate in candidate_roots:
            if candidate.joinpath(*CORPORATE_XLSX_CANONICAL_TEMPLATE_RELATIVE_PATH).exists():
                return candidate

        return candidate_roots[0]

    def generate(self, date_from, date_to):
        datasets = self.get_datasets(date_from, date_to)
        if not any(dataset.exists() for dataset in datasets.values()):
            raise EmptyCorporateXlsxDataError("No data for selected range.")

        workbook = load_workbook(self.template_path)
        self._populate_workbook(workbook, datasets)
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return CorporateXlsxArtifact(
            file_bytes=output.getvalue(),
            filename=f"corporate-qa-report-{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx",
        )

    def get_datasets(self, date_from, date_to):
        datasets = {}

        for dataset_config in CORPORATE_XLSX_EXPORT_CONFIG:
            model_class = self.MODEL_REGISTRY[dataset_config["model"]]
            queryset = model_class.objects.filter(**dataset_config["queryset_filters"])

            if dataset_config["date_field_type"] == "date":
                queryset = apply_datefield_date_range(
                    queryset,
                    dataset_config["date_field"],
                    date_from,
                    date_to,
                )
            else:
                queryset = apply_charfield_iso_date_range(
                    queryset,
                    dataset_config["date_field"],
                    date_from,
                    date_to,
                )

            datasets[dataset_config["dataset"]] = queryset

        return datasets

    def _resolve_template_path(self, template_path):
        if template_path:
            resolved_path = Path(template_path).resolve()
        else:
            resolved_path = self.project_root.joinpath(
                *CORPORATE_XLSX_CANONICAL_TEMPLATE_RELATIVE_PATH
            )

        placeholder_path = self.project_root.joinpath(
            *CORPORATE_XLSX_PLACEHOLDER_TEMPLATE_RELATIVE_PATH
        ).resolve()

        if resolved_path == placeholder_path:
            raise ValueError("Placeholder template is forbidden for corporate XLSX reports.")

        return resolved_path

    def _populate_workbook(self, workbook, datasets):
        for dataset_config in CORPORATE_XLSX_EXPORT_CONFIG:
            queryset = datasets[dataset_config["dataset"]]
            rows = self._queryset_to_rows(queryset, dataset_config)
            self._write_dataset_table(
                workbook,
                sheet_name=dataset_config["sheet_name"],
                table_name=dataset_config["table_name"],
                rows=rows,
            )

    @staticmethod
    def _queryset_to_rows(queryset, dataset_config):
        columns = dataset_config["columns"]
        row_builder = dataset_config.get("row_builder")

        if row_builder == "qc_fa":
            return CorporateXlsxReportService._build_qc_fa_rows(
                queryset=queryset,
                columns=columns,
                defect_columns=dataset_config.get("defect_columns", []),
            )

        if row_builder == "seconds_general":
            return CorporateXlsxReportService._build_seconds_general_rows(
                queryset=queryset,
                columns=columns,
            )

        if row_builder == "container":
            return CorporateXlsxReportService._build_container_rows(
                queryset=queryset,
                columns=columns,
                defect_columns=dataset_config.get("defect_columns", []),
            )

        values = queryset.order_by("pk").values_list(*columns)
        return [
            [CorporateXlsxReportService._normalize_cell_value(value) for value in row]
            for row in values
        ]

    @staticmethod
    def _build_qc_fa_rows(*, queryset, columns, defect_columns):
        rows = []
        defect_columns = set(defect_columns)
        inspections = queryset.select_related("color").prefetch_related(
            "inspection_defects__defect_type"
        ).order_by("pk")

        for inspection in inspections:
            defect_amounts = {
                defect.defect_type.name: defect.amount
                for defect in inspection.inspection_defects.all()
            }

            row = []
            for column in columns:
                if column == "color":
                    value = inspection.color.name if inspection.color_id else None
                elif column in defect_columns:
                    value = defect_amounts.get(column, 0)
                else:
                    value = getattr(inspection, column, None)

                row.append(CorporateXlsxReportService._normalize_cell_value(value))

            rows.append(row)

        return rows

    @staticmethod
    def _build_seconds_general_rows(*, queryset, columns):
        rows = []
        inspections = queryset.order_by("pk")

        for inspection in inspections:
            values = {
                column: CorporateXlsxReportService._normalize_cell_value(
                    getattr(inspection, column, None)
                )
                for column in columns
            }

            rows.append(
                [
                    values.get("date"),
                    values.get("week"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    values.get("total_de_tela"),
                    values.get("corrido_2"),
                    values.get("barre"),
                    values.get("otros_3"),
                    values.get("degradacion"),
                    values.get("bordados"),
                ]
            )

        return rows

    @staticmethod
    def _build_container_rows(*, queryset, columns, defect_columns):
        rows = []
        defect_columns = set(defect_columns)
        containers = queryset.prefetch_related("container_defects__defect_type").order_by("pk")

        for container in containers:
            defect_amounts = {
                defect.defect_type.name: defect.amount
                for defect in container.container_defects.all()
            }
            fallback_total_defects = sum(
                amount for name, amount in defect_amounts.items() if name != "total_defects"
            )

            row = []
            for column in columns:
                if column in defect_columns:
                    value = defect_amounts.get(column)
                    if column == "total_defects" and value is None:
                        value = fallback_total_defects
                    if value is None:
                        value = 0
                else:
                    value = getattr(container, column, None)

                row.append(CorporateXlsxReportService._normalize_cell_value(value))

            rows.append(row)

        return rows

    @staticmethod
    def _normalize_cell_value(value):
        if isinstance(value, (dt.date, dt.datetime)):
            return value.isoformat()
        return value

    def _write_dataset_table(self, workbook, *, sheet_name, table_name, rows):
        worksheet = workbook[sheet_name]
        table = self._get_table(worksheet, table_name)
        min_col, header_row, max_col, _ = range_boundaries(table.ref)
        prototype_row = header_row + 1

        self._clear_previous_table_body(
            worksheet,
            min_col=min_col,
            max_col=max_col,
            start_row=prototype_row,
            row_count=len(rows),
        )

        for row_offset, row_values in enumerate(rows):
            target_row = prototype_row + row_offset
            self._write_row_from_prototype(
                worksheet,
                target_row=target_row,
                prototype_row=prototype_row,
                min_col=min_col,
                max_col=max_col,
                row_values=row_values,
            )

        self._update_table_ref(
            table,
            min_col=min_col,
            max_col=max_col,
            header_row=header_row,
            row_count=len(rows),
        )

    @staticmethod
    def _get_table(worksheet, table_name):
        if table_name in worksheet.tables:
            return worksheet.tables[table_name]

        available_tables = list(worksheet.tables.keys())
        if len(available_tables) == 1:
            return worksheet.tables[available_tables[0]]

        raise KeyError(
            f"Table '{table_name}' not found in worksheet '{worksheet.title}'. "
            f"Available tables: {available_tables}"
        )

    @staticmethod
    def _clear_previous_table_body(worksheet, *, min_col, max_col, start_row, row_count):
        end_row = start_row if row_count == 0 else start_row + row_count - 1

        for row in range(start_row, end_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col).value = None

    def _write_row_from_prototype(
        self,
        worksheet,
        *,
        target_row,
        prototype_row,
        min_col,
        max_col,
        row_values,
    ):
        explicit_columns = len(row_values)

        for col_offset, col in enumerate(range(min_col, max_col + 1)):
            target_cell = worksheet.cell(row=target_row, column=col)
            source_cell = worksheet.cell(row=prototype_row, column=col)

            self._copy_cell_style(source_cell, target_cell)

            if col_offset < explicit_columns:
                target_cell.value = row_values[col_offset]
            else:
                target_cell.value = self._clone_formula_if_present(
                    source_cell=source_cell,
                    target_cell=target_cell,
                )

    @staticmethod
    def _copy_cell_style(source_cell, target_cell):
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.fill = copy(source_cell.fill)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)

    @staticmethod
    def _clone_formula_if_present(*, source_cell, target_cell):
        if source_cell.data_type != "f" or source_cell.value is None:
            return None

        try:
            return Translator(source_cell.value, origin=source_cell.coordinate).translate_formula(
                target_cell.coordinate
            )
        except Exception:
            return source_cell.value

    @staticmethod
    def _update_table_ref(table, *, min_col, max_col, header_row, row_count):
        last_row = header_row + row_count
        table.ref = (
            f"{get_column_letter(min_col)}{header_row}:"
            f"{get_column_letter(max_col)}{last_row}"
        )
