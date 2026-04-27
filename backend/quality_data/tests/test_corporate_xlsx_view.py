from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework import status as http_status
from rest_framework.test import APIClient

from quality_data.models import Color, QualityQcFa


class CorporateXlsxReportDateValidationTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.url = "/quality/reports/corporate-xlsx/"

    def test_missing_required_date_bounds_returns_400_with_field_errors(self):
        response = self.client.get(self.url, {"date_to": "2025-01-31"})

        self.assertEqual(response.status_code, http_status.HTTP_400_BAD_REQUEST)
        self.assertIn("date_from", response.data)

        response = self.client.get(self.url, {"date_from": "2025-01-01"})

        self.assertEqual(response.status_code, http_status.HTTP_400_BAD_REQUEST)
        self.assertIn("date_to", response.data)

    def test_invalid_date_format_returns_400_with_field_error(self):
        response = self.client.get(
            self.url,
            {"date_from": "01-01-2025", "date_to": "2025-01-31"},
        )

        self.assertEqual(response.status_code, http_status.HTTP_400_BAD_REQUEST)
        self.assertIn("date_from", response.data)

    def test_reversed_date_bounds_returns_400_with_explicit_order_error(self):
        response = self.client.get(
            self.url,
            {"date_from": "2025-02-01", "date_to": "2025-01-31"},
        )

        self.assertEqual(response.status_code, http_status.HTTP_400_BAD_REQUEST)
        self.assertIn("date_from", response.data)
        self.assertIn("date_to", response.data)


class CorporateXlsxReportContractTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.url = "/quality/reports/corporate-xlsx/"

    def test_valid_range_with_no_data_returns_422(self):
        response = self.client.get(
            self.url,
            {"date_from": "2025-01-01", "date_to": "2025-01-31"},
        )

        self.assertEqual(response.status_code, http_status.HTTP_422_UNPROCESSABLE_ENTITY)
        self.assertIn("error", response.data)

    def test_valid_range_with_data_returns_xlsx_attachment(self):
        color = Color.objects.create(name="Navy")
        QualityQcFa.objects.create(
            table_type="QFA",
            date_1="2025-01-15",
            week=3,
            customer="TestCustomer",
            team=1,
            coord="COORD-01",
            date_2="2025-01-15",
            po=1001,
            style="ST-001",
            batch=11,
            color=color,
            qty=120,
            seconds=2,
            accepted=118,
            rejected=2,
            sample=20,
            defects_total=2,
            aql=1.67,
            pass_or_fail="Pass",
        )
        QualityQcFa.objects.create(
            table_type="QFA",
            date_1="2025-03-15",
            week=11,
            customer="OutOfRangeCustomer",
            team=4,
            coord="COORD-OUT",
            date_2="2025-03-15",
            po=1999,
            style="ST-OUT",
            batch=19,
            color=color,
            qty=33,
            seconds=1,
            accepted=32,
            rejected=1,
            sample=20,
            defects_total=1,
            aql=3.03,
            pass_or_fail="Pass",
        )

        response = self.client.get(
            self.url,
            {"date_from": "2025-01-01", "date_to": "2025-01-31"},
        )

        self.assertEqual(response.status_code, http_status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        content_disposition = response["Content-Disposition"]
        self.assertIn("attachment;", content_disposition)
        self.assertIn(".xlsx", content_disposition)
        self.assertIn("2025-01-01", content_disposition)
        self.assertIn("2025-01-31", content_disposition)

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            [
                "Trims",
                "Packing Audit",
                "Sewing Endline 100% inspection",
                "Sewing In-Line",
                "QC FA Plant",
                "QC FA Customer",
                "SecondsA4",
                "Seconds General",
                "GraphxLine",
                "Container",
                "GeneralGraphics",
                "Thirds",
            ],
        )

        target_tables = {
            "QC FA Plant": "Table3",
            "QC FA Customer": "Table2",
            "SecondsA4": "Table15",
            "Seconds General": "Table1",
            "Container": "Table18",
        }
        for sheet_name, table_name in target_tables.items():
            self.assertIn(table_name, workbook[sheet_name].tables.keys())

        qfa_sheet = workbook["QC FA Plant"]
        qfa_table = qfa_sheet.tables["Table3"]
        self.assertEqual(qfa_table.ref, "A3:BP4")
        self.assertEqual(qfa_sheet["A4"].value, "2025-01-15")
        self.assertEqual(qfa_sheet["C4"].value, "TestCustomer")
        self.assertNotEqual(qfa_sheet["C4"].value, "OutOfRangeCustomer")

        workbook.close()
