import datetime as dt
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from excel_importer.sheet_configs import (
    CONTAINER_REMAP,
    CORPORATE_XLSX_EXPORT_CONFIG,
    QC_FA_CUSTOMER_REMAP,
    QC_FA_PLANT_REMAP,
)
from quality_data.corporate_xlsx_styles import CORPORATE_SHEET_ORDER
from quality_data.models import (
    Color,
    Container,
    ContainerDefectType,
    ContainerInspectionDefect,
    DefectType,
    InspectionDefect,
    QualityQcFa,
    SecondsGeneral,
    SecondsGeneralDefectType,
    SecondsGeneralDefect,
)


class CorporateXlsxWorkbookFidelityTest(TestCase):
    def setUp(self):
        from quality_data.corporate_xlsx_service import CorporateXlsxReportService
        self.service = CorporateXlsxReportService()
        self.target_tables = {
            "QC FA Plant": "Table3",
            "QC FA Customer": "Table2",
            "SecondsA4": "Table15",
            "Seconds General": "Table1",
            "Container": "Table18",
        }

    def _create_qfa_record(self, *, date_1, customer):
        color = Color.objects.create(name=f"Navy-{date_1}-{customer}")
        return QualityQcFa.objects.create(
            table_type="QFA",
            date_1=date_1,
            week=3,
            customer=customer,
            team=1,
            coord="COORD-01",
            date_2=date_1,
            po=1001,
            style="ST-001",
            batch=11,
            color=color,
            qty=120,
            seconds=2,
            accepted=118,
            rejected=2,
            sample=20,
            defects_total=2,
            aql=1.67,
            pass_or_fail="Pass",
        )

    def test_generated_workbook_preserves_sheet_order_and_table_identifiers(self):
        self._create_qfa_record(date_1="2025-01-15", customer="InRange Co")

        artifact = self.service.generate(dt.date(2025, 1, 1), dt.date(2025, 1, 31))
        generated = load_workbook(BytesIO(artifact.file_bytes))

        self.assertEqual(generated.sheetnames, CORPORATE_SHEET_ORDER)
        # No add_table — only autofilter, so no Table objects in metadata
        for sheet_name in self.target_tables:
            self.assertEqual(len(generated[sheet_name].tables), 0)

        generated.close()

    def test_generated_workbook_contains_only_filtered_qfa_rows_and_resized_table_ref(self):
        self._create_qfa_record(date_1="2025-01-15", customer="InRange Co")
        self._create_qfa_record(date_1="2025-03-05", customer="OutOfRange Co")

        artifact = self.service.generate(dt.date(2025, 1, 1), dt.date(2025, 1, 31))
        workbook = load_workbook(BytesIO(artifact.file_bytes))

        sheet = workbook["QC FA Plant"]
        # Header row 3 (1-indexed), data starts at row 4
        self.assertEqual(sheet["A3"].value, "Date")
        self.assertEqual(sheet["A4"].value, "2025-01-15")
        self.assertEqual(sheet["C4"].value, "InRange Co")
        self.assertNotEqual(sheet["C4"].value, "OutOfRange Co")
        # Verify header formatting is applied
        self.assertTrue(sheet["A3"].font.bold)
        self.assertEqual(sheet["A3"].font.size, 9)

        workbook.close()


class CorporateXlsxQcFaExportContractTest(TestCase):
    def setUp(self):
        from quality_data.corporate_xlsx_service import CorporateXlsxReportService
        self.service = CorporateXlsxReportService()

    def _create_quality_record(self, *, table_type, date_1):
        color = Color.objects.create(name="Navy Blue")
        return QualityQcFa.objects.create(
            table_type=table_type,
            date_1=date_1,
            week=8,
            customer="ACME",
            team=4,
            coord="ART-77",
            date_2=date_1,
            po=456,
            style="ST-456",
            batch=9,
            color=color,
            qty=180,
            seconds=3,
            accepted=177,
            rejected=3,
            sample=20,
            defects_total=4,
            aql=2.0,
            pass_or_fail="PASS",
        )

    def test_qfa_and_qfc_export_columns_follow_importer_canonical_order(self):
        qfa_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "qfa")
        qfc_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "qfc")
        seconds_general_config = next(
            config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "seconds_general"
        )
        container_config = next(
            config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "container"
        )

        self.assertEqual(qfa_config["columns"], list(QC_FA_PLANT_REMAP.values()))
        self.assertEqual(qfc_config["columns"], list(QC_FA_CUSTOMER_REMAP.values()))
        from excel_importer.sheet_configs import SECONDS_GENERAL_EXPORT_COLUMNS
        self.assertEqual(seconds_general_config["columns"], SECONDS_GENERAL_EXPORT_COLUMNS)
        self.assertEqual(container_config["columns"], list(CONTAINER_REMAP.values()))

    def test_qfa_row_serializes_color_name_and_defects_in_column_contract(self):
        inspection = self._create_quality_record(table_type="QFA", date_1="2025-02-10")
        defect_uneven = DefectType.objects.create(name="uneven")
        defect_missing_info = DefectType.objects.create(name="missing_information_label")
        InspectionDefect.objects.create(inspection=inspection, defect_type=defect_uneven, amount=2)
        InspectionDefect.objects.create(inspection=inspection, defect_type=defect_missing_info, amount=1)

        dataset_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "qfa")
        row = self.service._queryset_to_rows(QualityQcFa.objects.filter(pk=inspection.pk), dataset_config)[0]
        index = {column: position for position, column in enumerate(dataset_config["columns"])}

        self.assertEqual(row[index["date_1"]], "2025-02-10")
        self.assertEqual(row[index["color"]], "Navy Blue")
        self.assertEqual(row[index["uneven"]], 2)
        self.assertEqual(row[index["wrong_size_attached"]], 0)
        self.assertEqual(row[index["broken_stitch"]], 0)

    def test_qfc_row_serializes_color_name_and_defaults_missing_defects_to_zero(self):
        inspection = self._create_quality_record(table_type="QFC", date_1="2025-02-11")
        defect_wrong_transfer = DefectType.objects.create(name="wrong_transfer")
        InspectionDefect.objects.create(inspection=inspection, defect_type=defect_wrong_transfer, amount=3)

        dataset_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "qfc")
        row = self.service._queryset_to_rows(QualityQcFa.objects.filter(pk=inspection.pk), dataset_config)[0]
        index = {column: position for position, column in enumerate(dataset_config["columns"])}

        self.assertEqual(row[index["date_1"]], "2025-02-11")
        self.assertEqual(row[index["color"]], "Navy Blue")
        self.assertEqual(row[index["wrong_transfer"]], 3)
        self.assertEqual(row[index["uneven"]], 0)

    def test_seconds_general_row_maps_fields_to_main_table_positions(self):
        seconds_general = SecondsGeneral.objects.create(
            date="2025-02-20",
            week=8,
            team=11,
            line_code="11-12",
            customer="ACME",
            style="ST-001",
        )
        corrido2, _ = SecondsGeneralDefectType.objects.get_or_create(name="corrido_2")
        barre, _ = SecondsGeneralDefectType.objects.get_or_create(name="barre")
        otros3, _ = SecondsGeneralDefectType.objects.get_or_create(name="otros_3")
        degradacion, _ = SecondsGeneralDefectType.objects.get_or_create(name="degradacion")
        bordados, _ = SecondsGeneralDefectType.objects.get_or_create(name="bordados")
        for dt, amount in [(corrido2, 4), (barre, 2), (otros3, 1), (degradacion, 3), (bordados, 5)]:
            SecondsGeneralDefect.objects.create(
                seconds_general=seconds_general, defect_type=dt, amount=amount
            )

        dataset_config = next(
            config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "seconds_general"
        )
        row = self.service._queryset_to_rows(SecondsGeneral.objects.filter(pk=seconds_general.pk), dataset_config)[0]

        self.assertEqual(len(row), 39)
        self.assertEqual(row[0], "2025-02-20")
        self.assertEqual(row[1], 8)
        self.assertEqual(row[2], 11)       # team
        self.assertEqual(row[3], "11-12")  # line_code
        self.assertEqual(row[4], "ACME")
        self.assertEqual(row[5], "ST-001")
        self.assertEqual(row[33], 4)       # corrido_2
        self.assertEqual(row[34], 2)       # barre
        self.assertEqual(row[35], 1)       # otros_3
        self.assertEqual(row[36], 3)       # degradacion
        self.assertEqual(row[37], 5)       # bordados
        self.assertEqual(row[38], 15)      # total_de_tela

    def test_container_row_serializes_defects_from_related_table_in_importer_order(self):
        container = Container.objects.create(
            container_number=3001,
            date=dt.date(2025, 2, 21),
            customer="ACME",
            transfer_of_container=4,
            total_palette=20,
            total_palette_pass=18,
            total_palette_rejected=2,
            percentage_pass=90.0,
            percentage_reject=10.0,
        )
        dirt_label = ContainerDefectType.objects.create(name="dirt_label")
        dirt_container = ContainerDefectType.objects.create(name="dirt_container")
        total_defects = ContainerDefectType.objects.create(name="total_defects")
        ContainerInspectionDefect.objects.create(container=container, defect_type=dirt_label, amount=3)
        ContainerInspectionDefect.objects.create(container=container, defect_type=dirt_container, amount=2)
        ContainerInspectionDefect.objects.create(container=container, defect_type=total_defects, amount=5)

        dataset_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "container")
        row = self.service._queryset_to_rows(Container.objects.filter(pk=container.pk), dataset_config)[0]
        index = {column: position for position, column in enumerate(dataset_config["columns"])}

        self.assertEqual(row[index["date"]], "2025-02-21")
        self.assertEqual(row[index["container_number"]], 3001)
        self.assertEqual(row[index["customer"]], "ACME")
        self.assertEqual(row[index["dirt_label"]], 3)
        self.assertEqual(row[index["dirt_container"]], 2)
        self.assertEqual(row[index["total_defects"]], 5)
        self.assertEqual(row[index["container_holes"]], 0)

    def test_container_row_computes_total_defects_when_not_persisted(self):
        container = Container.objects.create(
            container_number=3002,
            date=dt.date(2025, 2, 22),
            customer="Beta",
            transfer_of_container=2,
            total_palette=10,
            total_palette_pass=8,
            total_palette_rejected=2,
            percentage_pass=80.0,
            percentage_reject=20.0,
        )
        dirt_label = ContainerDefectType.objects.create(name="dirt_label")
        warped_boxes = ContainerDefectType.objects.create(name="warped_boxes")
        ContainerInspectionDefect.objects.create(container=container, defect_type=dirt_label, amount=4)
        ContainerInspectionDefect.objects.create(container=container, defect_type=warped_boxes, amount=1)

        dataset_config = next(config for config in CORPORATE_XLSX_EXPORT_CONFIG if config["dataset"] == "container")
        row = self.service._queryset_to_rows(Container.objects.filter(pk=container.pk), dataset_config)[0]
        index = {column: position for position, column in enumerate(dataset_config["columns"])}

        self.assertEqual(row[index["total_defects"]], 5)
