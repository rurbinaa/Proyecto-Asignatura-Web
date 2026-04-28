"""
Programmatic Excel file generator for E2E tests.

Uses openpyxl to create a .xlsx file with 5 sheets matching the
Excel import schema. All data is deterministic.
"""

from pathlib import Path
from openpyxl import Workbook


def generate_test_excel(tmp_path) -> Path:
    """Generate a test .xlsx file with 5 sheets and return its Path.

    Sheets created:
    - QC FA Plant
    - QC FA Customer
    - SecondsA4
    - Seconds General
    - Container

    Args:
        tmp_path: pytest tmp_path fixture (pathlib.Path).

    Returns:
        Path to the generated .xlsx file.
    """
    wb = Workbook()

    # ── Sheet 1: QC FA Plant ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "QC FA Plant"
    ws1.append(["date_1", "week", "customer", "team", "coord", "po",
                 "style", "batch", "color", "qty", "seconds", "accepted",
                 "rejected", "sample", "defects_total", "aql",
                 "pass_or_fail", "sew_def", "fab_def"])
    ws1.append(["2025-01-10", 1, "CUST_A", 1, "JAVIER", 100,
                "N3165", 1, "Red", 100, 50, 95, 5, 100, 3, 2.5,
                "PASS", 2, 1])
    ws1.append(["2025-01-11", 1, "CUST_A", 2, "PEDRO", 101,
                "N4165", 1, "Blue", 100, 50, 90, 10, 100, 5, 2.5,
                "REJECT", 3, 2])
    ws1.append(["2025-01-12", 2, "CUST_B", 1, "JAVIER", 102,
                "N3165", 2, "Black", 100, 50, 85, 15, 100, 8, 2.5,
                "REJECT", 5, 3])

    # ── Sheet 2: QC FA Customer ───────────────────────────────────────────
    ws2 = wb.create_sheet("QC FA Customer")
    ws2.append(["date_1", "week", "customer", "team", "coord", "po",
                "style", "batch", "color", "qty", "seconds", "accepted",
                "rejected", "sample", "defects_total", "aql",
                "pass_or_fail", "sew_def", "fab_def"])
    ws2.append(["2025-01-10", 1, "CUST_C", 3, "MARIA", 200,
                "N5165", 1, "White", 80, 40, 78, 2, 80, 1, 1.5,
                "PASS", 1, 0])
    ws2.append(["2025-01-11", 1, "CUST_C", 3, "MARIA", 201,
                "N5165", 1, "Red", 80, 40, 76, 4, 80, 2, 1.5,
                "PASS", 1, 1])

    # ── Sheet 3: SecondsA4 ────────────────────────────────────────────────
    ws3 = wb.create_sheet("SecondsA4")
    ws3.append(["date_1", "week", "customer", "team", "coord", "po",
                "style", "batch", "color", "qty", "accepted", "rejected",
                "sample", "seconds"])
    ws3.append(["2025-01-10", 1, "CUST_A", 1, "JAVIER", 300,
                "N3165", 1, "Red", 50, 48, 2, 50, 20])
    ws3.append(["2025-01-11", 1, "CUST_A", 2, "PEDRO", 301,
                "N4165", 1, "Blue", 50, 45, 5, 50, 25])

    # ── Sheet 4: Seconds General ──────────────────────────────────────────
    ws4 = wb.create_sheet("Seconds General")
    ws4.append(["date_1", "week", "customer", "team", "coord", "po",
                "style", "batch", "color", "qty", "accepted", "rejected",
                "sample", "seconds"])
    ws4.append(["2025-01-10", 1, "CUST_B", 1, "JAVIER", 400,
                "N3165", 2, "Black", 30, 28, 2, 30, 15])

    # ── Sheet 5: Container ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Container")
    ws5.append(["container_id", "customer", "team", "coord", "po",
                "style", "batch", "color", "status", "qty", "accepted",
                "rejected", "sample", "defects_total"])
    ws5.append(["CONT-001", "CUST_A", 1, "JAVIER", 500,
                "N3165", 1, "Red", "Approved", 200, 195, 5, 200, 2])
    ws5.append(["CONT-002", "CUST_A", 2, "PEDRO", 501,
                "N4165", 1, "Blue", "Pending", 150, 148, 2, 150, 1])
    ws5.append(["CONT-003", "CUST_B", 1, "JAVIER", 502,
                "N3165", 2, "Black", "Approved", 180, 175, 5, 180, 3])

    filepath = tmp_path / "test_data.xlsx"
    wb.save(filepath)
    return filepath
